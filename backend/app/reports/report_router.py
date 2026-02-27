import csv
import io
import json
import re
from datetime import datetime
from typing import Tuple

from fastapi import APIRouter, HTTPException, Query
from fastapi.responses import FileResponse, Response
from openpyxl import Workbook

from app import storage, server_manager
from app.vuln_catalog import CATALOG, find_by_code, requires_manual_remediation
from .report_generator import (
    generate_analysis_global_pdf,
    generate_analysis_server_pdf,
    generate_remediation_global_pdf,
    generate_remediation_server_pdf,
)


def _is_server_connected(server_id: str) -> bool:
    """서버 SSH 연결 가능 여부 확인 (끊긴 서버 제외용)."""
    server = server_manager.get_server(server_id)
    if not server:
        return False
    try:
        from app.ssh_client import SSHClient
        with SSHClient(
            server["host"],
            server["port"],
            server["username"],
            server.get("password"),
            server.get("key_file"),
        ) as ssh:
            pass  # 연결 성공
        return True
    except Exception:
        return False
from .report_generator import (
    generate_analysis_global_pdf,
    generate_analysis_server_pdf,
    generate_remediation_global_pdf,
    generate_remediation_server_pdf,
)

router = APIRouter(prefix="/reports", tags=["Reports"])


# ================================
# 유틸: 저장된 분석 결과 기반 전체 진단 데이터 구성
# ================================
def _normalize_vuln_for_report(v: dict, hostname: str) -> dict:
    """main._normalize_vuln 과 유사하게 보고서용으로 정규화."""
    out = dict(v)
    if out.get("current_value") in (None, "") and out.get("detail"):
        out["current_value"] = out.get("detail") or ""
    out.setdefault("current_value", "")
    out.setdefault("expected_value", "")
    out.setdefault("details", [])
    out.setdefault("requires_manual_remediation", requires_manual_remediation(out.get("code") or out.get("check_id") or ""))
    # OS 정보 보존 (가능한 경우)
    out.setdefault("os_type", v.get("os_type"))
    out.setdefault("os_version", v.get("os_version"))
    out["hostname"] = hostname
    return out


def _format_detail_item(d) -> str:
    """details 항목을 읽기 쉬운 문자열로 변환 (dict → 깔끔한 포맷)."""
    if d is None:
        return ""
    if isinstance(d, dict):
        item = (d.get("점검항목") or "").strip()
        status = (d.get("상태") or "").strip()
        detail = (d.get("세부내용") or "").strip()
        if detail:
            # 세부내용 중심, 점검항목/상태 있으면 접두어로
            if item and status:
                return f"[{item}] {status}: {detail}"
            if item:
                return f"[{item}] {detail}"
            if status:
                return f"{status}: {detail}"
            return detail
        if item or status:
            parts = [f"점검항목: {item}" if item else "", f"상태: {status}" if status else ""]
            return " / ".join(p for p in parts if p)
        return json.dumps(d, ensure_ascii=False)
    return str(d).strip()


def _details_to_csv_cell(details) -> str:
    """details 리스트를 CSV 셀용 문자열로 (항목별 줄바꿈)."""
    if not details:
        return ""
    items = details if isinstance(details, list) else [details]
    lines = [_format_detail_item(d) for d in items if d]
    return "\n".join(lines)


def _analysis_to_csv(data: dict) -> str:
    """진단 보고서 데이터 → CSV 문자열 (UTF-8 BOM 포함, Excel 한글 인식용)."""
    vulns = data.get("vulnerabilities", [])
    is_global = "server_list" in data
    buf = io.StringIO()
    writer = csv.writer(buf)
    if is_global:
        headers = ["호스트명", "점검항목", "카테고리", "설명", "상태", "현재값", "권장값", "상세"]
    else:
        headers = ["점검항목", "카테고리", "설명", "상태", "현재값", "권장값", "상세"]
    writer.writerow(headers)
    status_ko = {"safe": "양호", "vulnerable": "취약", "manual": "조치 필요"}
    for v in vulns:
        code = v.get("code") or v.get("check_id") or ""
        cat = v.get("category") or ""
        desc = v.get("description") or v.get("name") or ""
        raw_status = (v.get("status") or "").strip().lower()
        status = "회귀" if v.get("regression") else (status_ko.get(raw_status) or raw_status or "")
        curr = v.get("current_value") or ""
        exp = v.get("expected_value") or ""
        details = v.get("details") or []
        detail_str = _details_to_csv_cell(details)
        row = [code, cat, desc, status, curr, exp, detail_str]
        if is_global:
            row.insert(0, v.get("hostname", ""))
        writer.writerow(row)
    return "\ufeff" + buf.getvalue()


def _split_remediation_details(details: list) -> Tuple[list, list]:
    """details를 '--- 조치 내역 ---' 기준으로 이전 값(진단 상세) / 조치 내역 으로 분리."""
    if not details:
        return [], []
    items = details if isinstance(details, list) else [details]
    split_idx = -1
    for i, d in enumerate(items):
        s = str(d).strip() if not isinstance(d, dict) else ""
        if s == "--- 조치 내역 ---":
            split_idx = i
            break
    if split_idx < 0:
        return items, []
    return items[:split_idx], items[split_idx + 1:]


def _extract_before_after_from_details(details: list) -> Tuple[str | None, str | None]:
    """
    details에서 기존 설정(조치 전) / 조치 후 설정 추출.
    - 기존: 진단 상세(prev)의 취약 내용 우선, 없으면 조치 내역의 '조치 전 상태'
    - 조치 후: 조치 내역의 '조치 후 상태' (실제 조치된 결과만 사용, 진단/점검 메시지 제외)
    """
    prev_items, remedy_items = _split_remediation_details(details or [])
    before_val: str | None = None
    after_val: str | None = None

    for d in remedy_items or []:
        if isinstance(d, dict):
            post = (d.get("조치 후 상태") or d.get("post_value") or "").strip()
            # 진단/점검 메시지("N개 ... 설정 미흡" 등) 제외, 실제 조치 결과만 사용
            if post and "설정 미흡" not in post:
                after_val = post
            pre = (d.get("조치 전 상태") or d.get("pre_value") or "").strip()
            if pre and not before_val:
                before_val = pre
        elif isinstance(d, str):
            s = d.strip()
            if s.startswith("조치 후:"):
                val = s[5:].strip()
                if val and "설정 미흡" not in val:
                    after_val = val
            elif s.startswith("조치 전:") and not before_val:
                before_val = s[5:].strip()
        if after_val and before_val:
            break

    for d in prev_items or []:
        if before_val:
            break
        if isinstance(d, dict):
            detail = (d.get("세부내용") or d.get("세부 내역") or d.get("detail") or "").strip()
            if detail and ("취약" in str(d.get("상태", "")) or "취약" in detail):
                before_val = detail
                break
        elif isinstance(d, str) and d.strip() and "취약" in d:
            before_val = d.strip()
            break

    return (before_val, after_val)


def _format_remediation_item(d) -> str:
    """조치 내역 항목을 읽기 쉬운 문자열로 변환."""
    if d is None:
        return ""
    if isinstance(d, dict):
        pre = (d.get("조치 전 상태") or d.get("pre_value") or "").strip()
        post = (d.get("조치 후 상태") or d.get("post_value") or "").strip()
        cmd = (d.get("조치 명령어") or "").strip()
        result = (d.get("조치 결과") or "").strip()
        sub = (d.get("세부 내역") or d.get("세부내용") or "").strip().replace("\\n", "\n")
        parts = []
        if pre:
            parts.append(f"조치 전: {pre}")
        if post:
            parts.append(f"조치 후: {post}")
        if cmd:
            parts.append(f"조치 명령: {cmd}")
        if result:
            parts.append(f"조치 결과: {result}")
        if sub:
            parts.append(sub)
        return "\n".join(parts) if parts else json.dumps(d, ensure_ascii=False)
    s = str(d).strip()
    return s


def _csv_to_rows(csv_content: str) -> list[list]:
    """CSV 문자열 → 행 리스트 (BOM 제거 후 파싱)."""
    s = csv_content.replace("\ufeff", "")
    reader = csv.reader(io.StringIO(s))
    return list(reader)


def _sanitize_sheet_name(name: str, used: set[str]) -> str:
    """Excel 시트명 유효화 (최대 31자, \ / ? * [ ] 제거, 중복 시 접미사)."""
    invalid = re.compile(r'[\\/:*?\[\]]')
    s = invalid.sub("_", str(name).strip())[:31]
    base = s
    idx = 0
    while s in used:
        idx += 1
        suffix = f"_{idx}"
        s = (base[: 31 - len(suffix)] + suffix)
    used.add(s)
    return s


def _remediation_to_csv(data: dict) -> str:
    """조치 보고서 데이터 → CSV 문자열 (이미지 형식: 점검영역, CODE, 점검항목, 위험도, 진단결과, 조치결과, 회귀, 이전 값, 조치 내역)."""
    vulns = data.get("vulnerabilities", [])
    buf = io.StringIO()
    writer = csv.writer(buf)
    # 이미지 참고: 점검영역, CODE, 점검항목, 위험도, 진단결과, 조치결과, 취약(현재설정)=이전값, 양호(조치내용)=조치내역
    headers = ["호스트명", "점검영역", "CODE", "점검항목", "위험도", "진단결과", "조치결과", "회귀", "이전 값(취약/현재설정)", "조치 내역"]
    writer.writerow(headers)

    sev_map = {"high": "H", "medium": "M", "low": "L"}

    for v in vulns:
        code = (v.get("check_id") or "").strip().upper()
        catalog = find_by_code(code) if code else None
        cat = catalog.category if catalog else ""
        desc = catalog.name if catalog else ""
        sev = sev_map.get((catalog.severity or "").lower(), "") if catalog else ""

        before_status = v.get("before_status") or ""
        after_status = v.get("after_status") or ""
        # 한글 표기
        status_ko = {"SAFE": "양호", "VULNERABLE": "취약", "MANUAL": "수동조치"}
        before_ko = status_ko.get(str(before_status).upper(), str(before_status))
        after_ko = status_ko.get(str(after_status).upper(), str(after_status))

        regression = v.get("regression")
        regress_str = "Y" if regression else ""

        details = v.get("details") or []
        prev_items, remedy_items = _split_remediation_details(details)

        # 이전 값: 진단 상세 (조치 전 상태) - 없으면 current_value 사용
        prev_str = _details_to_csv_cell(prev_items) if prev_items else (v.get("current_value") or "")

        # 조치 내역: "--- 조치 내역 ---" 이후
        remedy_lines = [_format_remediation_item(d) for d in remedy_items if d]
        remedy_str = "\n".join(remedy_lines) if remedy_lines else ""

        row = [
            v.get("hostname", ""),
            cat,
            code,
            desc,
            sev,
            before_ko,
            after_ko,
            regress_str,
            prev_str,
            remedy_str,
        ]
        writer.writerow(row)
    return "\ufeff" + buf.getvalue()


def _build_global_analysis_data():
    """
    - 등록된 서버 목록과 각 서버의 최신 분석 결과를 읽어서
      전체 진단 보고서에 필요한 데이터 구조로 변환.
    - summary.total_items 는 서버 수 × 카탈로그 항목 수로 계산.
    """
    all_servers = server_manager.list_servers()

    all_vulns: list[dict] = []
    server_list: list[dict] = []

    for server in all_servers:
        server_id = server["server_id"]
        analyses = storage.list_analyses_by_server(server_id)
        if not analyses:
            continue
        # 저장된 진단 결과가 있으면 보고서에 포함 (실시간 연결 확인 없음)

        # localdb.analysis_list_by_server 가 completed_at DESC 로 반환하므로
        # 인덱스 0 이 "가장 최신" 분석이다.
        latest = analyses[0]
        hostname = server.get("name", server.get("host", "unknown"))
        ip = server.get("host", "-")

        # OS / 버전 정보는 최신 분석의 취약점에서 보조적으로 추출
        vulns_raw = latest.get("vulnerabilities") or []
        sample_v = vulns_raw[0] if vulns_raw else {}
        os_type = sample_v.get("os_type") or server.get("server_type", "-")
        os_version = sample_v.get("os_version") or "-"

        # 서버 리스트 항목
        server_list.append(
            {
                "os": os_type,
                "osver": os_version,
                "ip": ip,
                "hostname": hostname,
            }
        )

        for v in latest.get("vulnerabilities", []):
            out = _normalize_vuln_for_report(v, hostname)
            all_vulns.append(out)

    if not all_vulns or not server_list:
        raise HTTPException(status_code=400, detail="진단 결과가 있는 서버가 없습니다.")

    # 전체 요약 정보 (result.json status 기준)
    safe_count = sum(1 for v in all_vulns if str(v.get("status", "")).lower() == "safe")
    vulnerable_count = sum(1 for v in all_vulns if str(v.get("status", "")).lower() == "vulnerable")
    manual_count = sum(1 for v in all_vulns if str(v.get("status", "")).lower() == "manual")

    # 전체 점검 항목 수 = 카탈로그 항목 수 × 서버 수
    total_items = len(CATALOG) * len(server_list)

    summary = {
        "total_targets": len(server_list),
        "total_items": total_items,
        "safe": safe_count,
        "vulnerable": vulnerable_count,
        "manual": manual_count,
        "action_required": vulnerable_count + manual_count,
    }

    return {
        "summary": summary,
        "vulnerabilities": all_vulns,
        "server_list": server_list,
    }


# ================================
# 1. 전체 점검 보고서
# ================================
@router.get("/analysis/global")
def analysis_global(format: str = Query(default="pdf", description="pdf, csv, json")):

    data = _build_global_analysis_data()

    if format == "csv":
        csv_content = _analysis_to_csv(data)
        return Response(
            content=csv_content.encode("utf-8"),
            media_type="text/csv; charset=utf-8",
            headers={"Content-Disposition": "attachment; filename=analysis_global_report.csv"},
        )
    if format == "json":
        out = {
            "generated_at": datetime.utcnow().isoformat() + "Z",
            "summary": data.get("summary", {}),
            "server_list": data.get("server_list", []),
            "vulnerabilities": data.get("vulnerabilities", []),
        }
        json_content = json.dumps(out, ensure_ascii=False, indent=2)
        return Response(
            content=json_content.encode("utf-8"),
            media_type="application/json; charset=utf-8",
            headers={"Content-Disposition": "attachment; filename=analysis_global_report.json"},
        )

    path = generate_analysis_global_pdf(data)
    return FileResponse(
        path,
        media_type="application/pdf",
        filename="analysis_global_report.pdf",
    )


# ================================
# 2. 특정 서버 점검
# ================================
@router.get("/analysis/server/{hostname}")
def analysis_server(hostname: str, format: str = Query(default="pdf", description="pdf, csv, json")):

    # hostname 기준으로 서버/분석 찾기
    servers = server_manager.list_servers()
    target = next(
        (s for s in servers if s.get("name") == hostname or s.get("host") == hostname),
        None,
    )
    if not target:
        raise HTTPException(status_code=404, detail="server_not_found")

    analyses = storage.list_analyses_by_server(target["server_id"])
    if not analyses:
        raise HTTPException(status_code=404, detail="analysis_not_found")

    latest = analyses[0]
    vulns_raw = latest.get("vulnerabilities") or []
    regression_codes = set(
        (c or "").strip().upper() for c in (latest.get("regression_codes") or []) if c
    )
    vulns = []
    for v in vulns_raw:
        out = dict(v)
        if out.get("current_value") in (None, "") and out.get("detail"):
            out["current_value"] = out.get("detail") or ""
        out.setdefault("current_value", "")
        out.setdefault("expected_value", "")
        out.setdefault("details", [])
        out.setdefault("requires_manual_remediation", requires_manual_remediation(out.get("code") or out.get("check_id") or ""))
        code = (out.get("check_id") or out.get("code") or "").strip().upper()
        out["regression"] = code in regression_codes
        vulns.append(out)

    # 서버 메타: OS 타입, OS 버전, IP, 호스트네임
    sample_v = vulns_raw[0] if vulns_raw else {}
    server_meta = {
        "os_type": sample_v.get("os_type") or target.get("server_type", "-"),
        "os_version": sample_v.get("os_version", "-"),
        "ip": target.get("host", "-"),
        "hostname": hostname,
    }

    # 요약 집계: result.json status 기준 (양호/취약/기타)
    safe_count = sum(1 for v in vulns if str(v.get("status", "")).lower() == "safe")
    vulnerable_count = sum(1 for v in vulns if str(v.get("status", "")).lower() == "vulnerable")
    manual_count = sum(1 for v in vulns if str(v.get("status", "")).lower() == "manual")

    total_items = len(vulns)

    data = {
        "hostname": hostname,
        "server_meta": server_meta,
        "vulnerabilities": vulns,
        "summary": {
            "total_items": total_items,
            "safe": safe_count,
            "vulnerable": vulnerable_count,
            "manual": manual_count,
            "action_required": vulnerable_count + manual_count,
        },
    }

    if format == "csv":
        csv_content = _analysis_to_csv(data)
        return Response(
            content=csv_content.encode("utf-8"),
            media_type="text/csv; charset=utf-8",
            headers={"Content-Disposition": f"attachment; filename=analysis_{hostname}.csv"},
        )
    if format == "json":
        out = {
            "generated_at": datetime.utcnow().isoformat() + "Z",
            "hostname": hostname,
            "server_meta": server_meta,
            "summary": data.get("summary", {}),
            "vulnerabilities": vulns,
        }
        json_content = json.dumps(out, ensure_ascii=False, indent=2)
        return Response(
            content=json_content.encode("utf-8"),
            media_type="application/json; charset=utf-8",
            headers={"Content-Disposition": f"attachment; filename=analysis_{hostname}.json"},
        )

    path = generate_analysis_server_pdf(data)
    return FileResponse(
        path,
        media_type="application/pdf",
        filename=f"analysis_{hostname}.pdf",
    )


# ================================
# 조치 항목별 스냅샷에서 관련 설정 라인 추출 (증거/코드)
# ================================
# check_id → 스냅샷에서 찾을 때 사용할 키 접두사 목록 (storage._apply_snapshot_fix와 동기화)
_CODE_SNAPSHOT_PREFIXES: dict[str, list[str]] = {
    "U-01": ["PermitRootLogin", "PasswordAuthentication"],
    "U-02": ["PASS_MIN_LEN"],
    "U-47": ["PASS_MIN_DAYS"],
    "U-18": ["PermitRootLogin", "PasswordAuthentication"],
}


def _snapshot_lines_for_code(code: str, snapshot: list[str]) -> list[str]:
    """해당 항목(check_id)과 관련된 스냅샷 라인만 추출 (문제/증거용)."""
    if not snapshot:
        return []
    code = (code or "").strip().upper()
    prefixes = _CODE_SNAPSHOT_PREFIXES.get(code)
    if not prefixes:
        prefixes = [f"FIXED_{code}", code.replace("-", "_")]
    lines = []
    for line in snapshot:
        s = line.strip()
        if not s:
            continue
        if any(s.startswith(p) for p in prefixes):
            lines.append(line)
        elif code in s or code.replace("-", "") in s.replace("-", ""):
            lines.append(line)
    return lines


# ================================
# 조치 보고서용 데이터 구성 (before/after 비교)
# ================================
def _build_remediation_data(hostname_filter: str | None = None):
    """
    - 각 서버별로 가장 오래된 분석(before)과 최신 분석(after)을 비교해
      조치 보고서용 vulnerabilities, summary, server_list 생성.
    - hostname_filter 가 있으면 해당 서버만 포함.
    """
    all_servers = server_manager.list_servers()
    if hostname_filter:
        all_servers = [
            s for s in all_servers
            if s.get("name") == hostname_filter or s.get("host") == hostname_filter
        ]
        if not all_servers:
            raise HTTPException(status_code=404, detail="server_not_found")

    all_vulns: list[dict] = []
    server_list: list[dict] = []

    for server in all_servers:
        server_id = server["server_id"]
        analyses = storage.list_analyses_by_server(server_id)
        if not analyses:
            continue
        # 저장된 진단/조치 결과가 있으면 보고서에 포함

        # completed_at DESC → [0]=최신, [1]=그 이전, ... 비교 대상을 "가장 최근 두 건"으로 해서 보고서가 새 진단 시 갱신되도록 함
        after_analysis = analyses[0]
        before_analysis = analyses[1] if len(analyses) >= 2 else analyses[0]
        is_same_analysis = (before_analysis.get("analysis_id") == after_analysis.get("analysis_id"))

        hostname = server.get("name", server.get("host", "unknown"))
        ip = server.get("host", "-")
        vulns_before = {v.get("code") or v.get("check_id"): v for v in (before_analysis.get("vulnerabilities") or [])}
        vulns_after = {v.get("code") or v.get("check_id"): v for v in (after_analysis.get("vulnerabilities") or [])}
        all_codes = set(vulns_before) | set(vulns_after)

        sample_v = list(vulns_after.values())[0] if vulns_after else (list(vulns_before.values())[0] if vulns_before else {})
        os_type = sample_v.get("os_type") or server.get("server_type", "-")
        os_version = sample_v.get("os_version") or "-"

        server_list.append({
            "os": os_type,
            "osver": os_version,
            "ip": ip,
            "hostname": hostname,
        })

        for code in all_codes:
            b = vulns_before.get(code) or {}
            a = vulns_after.get(code) or {}
            before_status = (b.get("status") or "").strip().upper() or None
            after_status = (a.get("status") or "").strip().upper() or None
            if not before_status:
                before_status = "VULNERABLE"
            if not after_status:
                after_status = before_status
            regression = (before_status == "SAFE" and after_status == "VULNERABLE") or (before_status == "SAFE" and after_status == "MANUAL")

            curr_val = b.get("current_value") or b.get("detail") or ""
            exp_val = a.get("current_value") or a.get("expected_value") or a.get("detail") or ""
            details = a.get("details") or b.get("details") or []
            if not details and (a.get("detail") or b.get("detail")):
                details = [a.get("detail") or b.get("detail")]

            # 조치 로직은 vulnerable/manual 항목만 건드림. safe -> safe 는 보고서에서 제외
            if before_status == "SAFE" and after_status == "SAFE":
                continue

            # 조치 전/후 스냅샷에서 해당 항목 관련 설정 라인 추출 (증거/코드)
            before_snap = before_analysis.get("snapshot") or []
            after_snap = after_analysis.get("snapshot") or []
            before_evidence = _snapshot_lines_for_code(code, before_snap)
            after_evidence = _snapshot_lines_for_code(code, after_snap)

            # 스냅샷에 증거가 없으면 details(조치 내역)에서 조치 전/후 상태 추출
            extracted_before, extracted_after = _extract_before_after_from_details(details)
            if not before_evidence and extracted_before:
                before_evidence = [extracted_before]
            if not after_evidence and extracted_after:
                after_evidence = [extracted_after]
            # 기존설정 / 조치후설정을 조치 내역에서 추출한 값으로 보완 (같은 내용이 둘 다 나오는 문제 방지)
            if extracted_before:
                curr_val = extracted_before
            if extracted_after:
                exp_val = extracted_after

            _, remedy_items = _split_remediation_details(details)
            has_remediation_detail = bool(remedy_items)

            # 조치가 같은 analysis 내에서 이뤄졌을 때: before/after가 동일해 둘 다 SAFE/FIXED로 나옴.
            # details에 조치 내역이 있으면 → 조치 전은 취약이었음으로 보정 (요약/개선 개수 반영)
            effective_before = before_status if before_status else "VULNERABLE"
            if is_same_analysis and has_remediation_detail and str(after_status).upper() in ("SAFE", "FIXED"):
                if str(before_status).upper() in ("SAFE", "FIXED"):
                    effective_before = "VULNERABLE"

            all_vulns.append({
                "check_id": (code or "").strip().upper(),
                "hostname": hostname,
                "before_status": effective_before,
                "after_status": after_status if after_status else effective_before,
                "regression": regression,
                "current_value": curr_val,
                "expected_value": exp_val,
                "after_value": exp_val,
                "details": details if isinstance(details, list) else [details],
                "before_evidence": before_evidence,
                "after_evidence": after_evidence,
                "has_remediation_detail": has_remediation_detail,
            })

    if not all_vulns or not server_list:
        raise HTTPException(status_code=400, detail="연결된 조치 서버가 없습니다. 끊긴 서버는 보고서에서 제외됩니다.")

    before_vulnerable = sum(1 for v in all_vulns if str(v.get("before_status", "")).upper() in ("VULNERABLE", "MANUAL"))
    after_vulnerable = sum(1 for v in all_vulns if str(v.get("after_status", "")).upper() in ("VULNERABLE", "MANUAL"))
    after_safe_or_fixed = ("SAFE", "FIXED")
    improved = sum(1 for v in all_vulns if str(v.get("before_status", "")).upper() in ("VULNERABLE", "MANUAL") and str(v.get("after_status", "")).upper() in after_safe_or_fixed)
    regression_count = sum(1 for v in all_vulns if v.get("regression"))
    total_items = len(all_vulns)
    improvement_rate = round((improved / before_vulnerable * 100), 1) if before_vulnerable else 0

    summary = {
        "total_targets": len(server_list),
        "total_items": total_items,
        "before_vulnerable": before_vulnerable,
        "after_vulnerable": after_vulnerable,
        "improved": improved,
        "improvement_rate": improvement_rate,
        "regression_count": regression_count,
        "total": before_vulnerable,
        "unresolved": after_vulnerable,
        "regressions": regression_count,
    }

    return {
        "summary": summary,
        "vulnerabilities": all_vulns,
        "server_list": server_list,
        "hostname": hostname_filter,
    }


# ================================
# 3. 전체 조치 보고서
# ================================
@router.get("/remediation/global")
def remediation_global(format: str = Query(default="pdf", description="pdf, csv, json")):

    data = _build_remediation_data()

    if format == "csv":
        csv_content = _remediation_to_csv(data)
        return Response(
            content=csv_content.encode("utf-8"),
            media_type="text/csv; charset=utf-8",
            headers={"Content-Disposition": "attachment; filename=remediation_global_report.csv"},
        )
    if format == "json":
        out = {
            "generated_at": datetime.utcnow().isoformat() + "Z",
            "summary": data.get("summary", {}),
            "server_list": data.get("server_list", []),
            "vulnerabilities": data.get("vulnerabilities", []),
        }
        json_content = json.dumps(out, ensure_ascii=False, indent=2)
        return Response(
            content=json_content.encode("utf-8"),
            media_type="application/json; charset=utf-8",
            headers={"Content-Disposition": "attachment; filename=remediation_global_report.json"},
        )

    path = generate_remediation_global_pdf(data)
    return FileResponse(
        path,
        media_type="application/pdf",
        filename="remediation_global_report.pdf",
    )


# ================================
# 4. 특정 서버 조치
# ================================
@router.get("/remediation/server/{hostname}")
def remediation_server(hostname: str, format: str = Query(default="pdf", description="pdf, csv, json")):

    data = _build_remediation_data(hostname_filter=hostname)
    data["hostname"] = hostname
    # 개별 보고서용 메타: 해당 서버 정보
    server_list = data.get("server_list") or []
    server_meta = next((s for s in server_list if s.get("hostname") == hostname), {})

    if format == "csv":
        csv_content = _remediation_to_csv(data)
        return Response(
            content=csv_content.encode("utf-8"),
            media_type="text/csv; charset=utf-8",
            headers={"Content-Disposition": f"attachment; filename=remediation_{hostname}.csv"},
        )
    if format == "json":
        out = {
            "generated_at": datetime.utcnow().isoformat() + "Z",
            "hostname": hostname,
            "server_meta": server_meta,
            "summary": data.get("summary", {}),
            "vulnerabilities": data.get("vulnerabilities", []),
        }
        json_content = json.dumps(out, ensure_ascii=False, indent=2)
        return Response(
            content=json_content.encode("utf-8"),
            media_type="application/json; charset=utf-8",
            headers={"Content-Disposition": f"attachment; filename=remediation_{hostname}.json"},
        )

    data["server_meta"] = server_meta
    path = generate_remediation_server_pdf(data)
    return FileResponse(
        path,
        media_type="application/pdf",
        filename=f"AUTOISMS 개별 조치 보고서_{hostname}.pdf",
    )


# ================================
# 5. 개별 진단 보고서 전체 → Excel (시트별 타겟)
# ================================
@router.get("/analysis/individuals/excel")
def analysis_individuals_excel():
    """
    진단된 모든 서버의 개별 진단 보고서를 하나의 Excel 파일로 생성.
    각 시트 = 한 서버의 진단 CSV 데이터.
    """
    try:
        global_data = _build_global_analysis_data()
    except HTTPException:
        raise

    server_list = global_data.get("server_list") or []
    all_vulns = global_data.get("vulnerabilities") or []

    if not server_list:
        raise HTTPException(status_code=400, detail="진단 결과가 있는 서버가 없습니다.")

    wb = Workbook()
    used_sheet_names: set[str] = set()

    for server in server_list:
        hostname = server.get("hostname", "unknown")
        vulns = [v for v in all_vulns if v.get("hostname") == hostname]
        safe_count = sum(1 for v in vulns if str(v.get("status", "")).lower() == "safe")
        vulnerable_count = sum(1 for v in vulns if str(v.get("status", "")).lower() == "vulnerable")
        manual_count = sum(1 for v in vulns if str(v.get("status", "")).lower() == "manual")

        data = {
            "hostname": hostname,
            "server_meta": {
                "os_type": server.get("os", "-"),
                "os_version": server.get("osver", "-"),
                "ip": server.get("ip", "-"),
                "hostname": hostname,
            },
            "vulnerabilities": vulns,
            "summary": {
                "total_items": len(vulns),
                "safe": safe_count,
                "vulnerable": vulnerable_count,
                "manual": manual_count,
                "action_required": vulnerable_count + manual_count,
            },
        }
        csv_content = _analysis_to_csv(data)
        rows = _csv_to_rows(csv_content)
        sheet_name = _sanitize_sheet_name(hostname, used_sheet_names)
        ws = wb.create_sheet(title=sheet_name)
        for r_idx, row in enumerate(rows, start=1):
            for c_idx, val in enumerate(row, start=1):
                ws.cell(row=r_idx, column=c_idx, value=val)

    # 첫 번째 시트(openpyxl 기본 Sheet) 제거 후 시트 순서 정리
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    return Response(
        content=buf.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=analysis_individuals.xlsx"},
    )


# ================================
# 6. 개별 조치 보고서 전체 → Excel (시트별 타겟)
# ================================
@router.get("/remediation/individuals/excel")
def remediation_individuals_excel():
    """
    조치된 모든 서버의 개별 조치 보고서를 하나의 Excel 파일로 생성.
    각 시트 = 한 서버의 조치 CSV 데이터.
    """
    try:
        global_data = _build_remediation_data()
    except HTTPException:
        raise

    server_list = global_data.get("server_list") or []
    if not server_list:
        raise HTTPException(status_code=400, detail="조치 결과가 있는 서버가 없습니다.")

    wb = Workbook()
    used_sheet_names: set[str] = set()

    for server in server_list:
        hostname = server.get("hostname", "unknown")
        data = _build_remediation_data(hostname_filter=hostname)
        data["hostname"] = hostname
        csv_content = _remediation_to_csv(data)
        rows = _csv_to_rows(csv_content)
        sheet_name = _sanitize_sheet_name(hostname, used_sheet_names)
        ws = wb.create_sheet(title=sheet_name)
        for r_idx, row in enumerate(rows, start=1):
            for c_idx, val in enumerate(row, start=1):
                ws.cell(row=r_idx, column=c_idx, value=val)

    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    return Response(
        content=buf.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=remediation_individuals.xlsx"},
    )